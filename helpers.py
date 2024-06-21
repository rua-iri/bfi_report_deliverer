from hashlib import md5 as md5_hash
import json
import os
from time import time as unix_timestamp
import openpyxl
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import constants
import pdfkit
import sqlite3
import jinja2
import requests
import dotenv
from classes import Film
from mapping import (
    RANK,
    TITLE,
    ORIGIN_COUNTRY,
    WEEKEND_GROSS,
    DISTRIBUTOR,
    WEEKLY_CHANGE,
    WEEKS_ON_RELEASE,
    CINEMA_NUMBER,
    SITE_AVERAGE,
    TOTAL_GROSS,
)


top_15_row: list = [3, 17]
other_uk_row: list = [22]

dotenv.load_dotenv()


con = sqlite3.connect("bfi_report.db")
con.row_factory = sqlite3.Row

jinja_environment = jinja2.Environment(loader=jinja2.FileSystemLoader("templates"))


def get_film_data(film: Film) -> dict:
    response = requests.get(
        constants.TMDB_SEARCH_API_URL.format(query=film.title.split("(")[0]),
        headers={
            "accept": "application/json",
            "Authorization": os.environ.get("TMBD_API_KEY"),
        },
    )

    film_data = json.loads(response.text)

    if film_data["total_results"] == 0:
        return {}

    first_result = film_data["results"][0]
    film_id = film_data["results"][0]["id"]

    print(first_result)

    response = requests.get(
        constants.TMDB_DETAILS_API_URL.format(id=film_id),
        headers={
            "accept": "application/json",
            "Authorization": os.environ.get("TMBD_API_KEY"),
        },
    )

    film_data = json.loads(response.text)

    return {"poster": film_data["poster_path"], "imdb_id": film_data["imdb_id"]}


def parse_films(film_group: str) -> list:
    film_list = []

    sheet = openpyxl.load_workbook(
        filename=constants.FILE_DOWNLOAD_LOCATION, read_only=True, data_only=True
    ).active

    if film_group == "top_15":
        MIN_ROW = top_15_row[0]
        MAX_ROW = top_15_row[1]
    elif film_group == "other_uk":
        MIN_ROW = other_uk_row[0]
        MAX_ROW = sheet.max_row
    elif film_group == "other_new":
        MIN_ROW = other_uk_row[1] + 3
        MAX_ROW = sheet.max_row
    else:
        raise Exception("Invalid Film Group")

    for index, row in enumerate(
        sheet.iter_rows(
            min_row=MIN_ROW,
            max_row=MAX_ROW,
            min_col=constants.MIN_COL,
            max_col=constants.MAX_COL,
            values_only=True,
        )
    ):

        if row[RANK] == None:
            if film_group == "other_uk":
                other_uk_row.append(index + other_uk_row[0])
            break

        film = Film(
            rank=row[RANK],
            title=row[TITLE],
            origin_country=row[ORIGIN_COUNTRY],
            weekend_gross=row[WEEKEND_GROSS],
            distributor=row[DISTRIBUTOR],
            weekly_change=row[WEEKLY_CHANGE],
            weeks_on_release=row[WEEKS_ON_RELEASE],
            cinema_number=row[CINEMA_NUMBER],
            site_average=row[SITE_AVERAGE],
            total_gross=row[TOTAL_GROSS],
        )
        film_data = get_film_data(film)
        film.set_film_data(film_data)
        film_list.append(film)

    return film_list


def generate_html_report(film_list: list, weekend_date: str) -> None:
    """
    Generate the new html report using the html templates
    """
    base_html = jinja_environment.get_template(constants.BASE_HTML_TEMPLATE)
    card_html = jinja_environment.get_template(constants.CARD_TEMPLATE)

    complete_card_html = ""

    # generate table rows and append them to complete table string
    for film in film_list:
        card = card_html.render(film.__dict__)
        complete_card_html += card

    # write content to html
    with open(constants.HTML_REPORT_LOCATION, "w") as file:
        file.write(
            base_html.render(
                top_15_contents=complete_card_html, weekend_date=weekend_date
            )
        )


def generate_pdf_report() -> None:
    """
    Generate the new pdf report using the html template
    """
    pdfkit.from_file(
        input=constants.HTML_REPORT_LOCATION, output_path=constants.PDF_REPORT_LOCATION
    )


def get_subscribers():
    """
    Query database for list of all currently active subscribers
    """
    cursor = con.cursor()
    res = cursor.execute(
        constants.SELECT_USERS_QUERY,
    )
    user_list = res.fetchall()

    return user_list


def generate_email_body(user_name: str, week_number: str) -> str:
    """
    Generate the body of the email to be sent to subscribers
    """

    email_template = jinja_environment.get_template("email.html")
    return email_template.render(user_name=user_name, week_number=week_number)


def create_db():
    """
    Create users table in database (should only run on initialising repository)
    """
    cursor = con.cursor()
    res = cursor.execute(constants.CREATE_TABLE_QUERY)
    res.fetchall()


def gen_file_hash() -> str:
    """Generate a hash of the excel report file downloaded

    Returns:
        str: An md5 hash of the file
    """

    h_lib = md5_hash()

    with open(constants.FILE_DOWNLOAD_LOCATION, "rb") as file:
        chunk = 0
        while chunk != b"":
            chunk = file.read(1024)
            h_lib.update(chunk)

    return h_lib.hexdigest()


def is_file_new(file_hash: str) -> bool:
    """Check whether the file downloaded has been processed before and insert
    the new file's hash into the database

    Args:
        file_hash (str): An md5 hash of the file from gen_file_hash()

    Returns:
        bool: Whether the file is new or not
    """

    cursor = con.cursor()
    result = cursor.execute(constants.SELECT_FILES_QUERY, (file_hash,)).fetchall()
    print(result)
    print(not result)

    timestamp = int(unix_timestamp())

    cursor.execute(constants.INSERT_FILE_QUERY, (file_hash, timestamp))
    con.commit()

    if len(result) == 0:
        return True
    else:
        return False


if __name__ == "__main__":
    # is_file_new("67089608790baa07d928c8d5f51b5c28")
    get_film_data("The Fall Guy")
