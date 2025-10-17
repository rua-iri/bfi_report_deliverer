import json
import openpyxl
import pdfkit
import sqlite3
import jinja2
import requests
from os import environ, makedirs
from os.path import dirname
from hashlib import md5 as md5_hash
from time import time as unix_timestamp
from dotenv import load_dotenv

from . import constants
from .queries import (SELECT_USERS_QUERY,
                      SELECT_FILES_QUERY,
                      INSERT_FILE_QUERY
                      )
from .classes import Film
from .mapping import (
    RANK,
    TITLE,
    ORIGIN_COUNTRY,
    WEEKEND_GROSS,
    DISTRIBUTOR,
    WEEKLY_CHANGE,
    WEEKS_ON_RELEASE,
    CINEMA_NUMBER,
    SITE_AVERAGE,
    TOTAL_GROSS,
)


other_uk_row: list = [22]

load_dotenv()


con = sqlite3.connect("bfi_report.db")
con.row_factory = sqlite3.Row

jinja_environment = jinja2.Environment(
    loader=jinja2.FileSystemLoader("templates"))


def create_dir(file_path):
    makedirs(
        dirname(file_path),
        exist_ok=True
    )


def fetch_film_data(film: Film) -> dict:
    """Fetch film data from TMDB api

    Args:
        film (Film): an instance of the Film class

    Returns:
        dict: a dictionary containing detailed information about the film
    """
    response = requests.get(
        constants.TMDB_SEARCH_API_URL.format(query=film.title.split("(")[0]),
        headers={
            "accept": "application/json",
            "Authorization": f"Bearer {environ.get("TMBD_API_KEY")}",
        },
    )

    film_data = json.loads(response.text)

    if film_data["total_results"] == 0:
        return {}

    film_id = film_data["results"][0]["id"]

    response = requests.get(
        constants.TMDB_DETAILS_API_URL.format(id=film_id),
        headers={
            "accept": "application/json",
            "Authorization": f"Bearer {environ.get("TMBD_API_KEY")}",
        },
    )

    film_data = json.loads(response.text)

    return {
        "poster": film_data["poster_path"],
        "imdb_id": film_data["imdb_id"]
    }


def parse_films(
        film_group: str,
        filename: str = constants.FILE_DOWNLOAD_LOCATION
) -> list:
    """Parse the spreadsheet to get top films

    Args:
        film_group (str): The group of films being parsed for

    Raises:
        Exception: exception if the film group passed is invalid

    Returns:
        list: a list of top films for a given category
    """
    film_list = []

    sheet = openpyxl.load_workbook(
        filename=filename,
        read_only=True,
        data_only=True
    ).active

    if film_group == "top_15":
        MIN_ROW = constants.TOP_15_MIN
        MAX_ROW = constants.TOP_15_MAX
    elif film_group == "other_uk":
        MIN_ROW = other_uk_row[0]
        MAX_ROW = sheet.max_row
    elif film_group == "other_new":
        MIN_ROW = other_uk_row[1] + 3
        MAX_ROW = sheet.max_row
    else:
        raise Exception("Invalid Film Group")

    for index, row in enumerate(
        sheet.iter_rows(
            min_row=MIN_ROW,
            max_row=MAX_ROW,
            min_col=constants.MIN_COL,
            max_col=constants.MAX_COL,
            values_only=True,
        )
    ):

        if row[RANK] is None:
            if film_group == "other_uk":
                other_uk_row.append(index + other_uk_row[0])
            break

        film = Film(
            rank=row[RANK],
            title=row[TITLE],
            origin_country=row[ORIGIN_COUNTRY],
            weekend_gross=row[WEEKEND_GROSS],
            distributor=row[DISTRIBUTOR],
            weekly_change=row[WEEKLY_CHANGE],
            weeks_on_release=row[WEEKS_ON_RELEASE],
            cinema_number=row[CINEMA_NUMBER],
            site_average=row[SITE_AVERAGE],
            total_gross=row[TOTAL_GROSS],
        )
        film_data = fetch_film_data(film)
        film.set_film_data(film_data)
        film_list.append(film)

    return film_list


def render_html(film_list: list) -> str:
    card_html = jinja_environment.get_template(constants.CARD_TEMPLATE)
    film_list_contents: str = ""
    page_count: int = 0
    film_on_page_count: int = 0

    # cards and append them to complete table string
    for film in film_list:
        card = card_html.render(film.__dict__)
        film_list_contents += card

        film_on_page_count += 1
        if page_count > 0 and film_on_page_count == 4:
            film_list_contents += constants.HTML_PAGE_BREAK
            film_on_page_count = 0
            page_count += 1
        elif page_count == 0 and film_on_page_count == 3:
            film_list_contents += constants.HTML_PAGE_BREAK
            film_on_page_count = 0
            page_count += 1

    film_list_contents += constants.HTML_PAGE_BREAK

    return film_list_contents


def generate_html_report(
    top_15_film_list: list,
    other_uk_film_list: list,
    other_new_film_list: list,
    weekend_date: str,
) -> None:
    """Generate the new html report using the html templates

    Args:
        top_15_film_list (list): List of top 15 films
        other_uk_film_list (list): List of other top uk films
        other_new_film_list (list): List of other top new films
        weekend_date (str): The date of the previous weekend
    """
    base_html = jinja_environment.get_template(constants.BASE_HTML_TEMPLATE)

    top_15_contents = render_html(top_15_film_list)
    other_uk_contents = render_html(other_uk_film_list)
    other_new_contents = render_html(other_new_film_list)

    # write content to html
    create_dir(constants.HTML_REPORT_LOCATION)
    with open(constants.HTML_REPORT_LOCATION, "w") as file:
        file.write(
            base_html.render(
                top_15_contents=top_15_contents,
                other_uk_contents=other_uk_contents,
                other_new_contents=other_new_contents,
                weekend_date=weekend_date,
            )
        )


def generate_pdf_report() -> None:
    """Generate the new pdf report using the html template
    """
    try:
        create_dir(constants.HTML_REPORT_LOCATION)
        pdfkit.from_file(
            input=constants.HTML_REPORT_LOCATION,
            output_path=constants.PDF_REPORT_LOCATION
        )

    except Exception as e:
        raise e


def get_subscribers() -> list:
    """Query database for list of all currently active subscribers

    Returns:
        list: list of all currently subscribed users
    """
    cursor = con.cursor()
    res = cursor.execute(
        SELECT_USERS_QUERY,
    )
    user_list = res.fetchall()

    return user_list


def generate_email_body(user_name: str, week_number: str) -> str:
    """Generate the body of the email to be sent to subscribers

    Args:
        user_name (str): The user to whom the email will be sent
        week_number (str): The date of the previous weekend

    Returns:
        str: html content of the email
    """

    email_template = jinja_environment.get_template("email.html")
    return email_template.render(user_name=user_name, week_number=week_number)


def gen_file_hash(original_filename: str) -> str:
    """Generate a hash of the excel report file downloaded

    Returns:
        str: An md5 hash of the file
    """

    h_lib = md5_hash()

    with open(original_filename, "rb") as file:
        chunk = 0
        while chunk != b"":
            chunk = file.read(1024)
            h_lib.update(chunk)

    return h_lib.hexdigest()


def is_file_new(file_hash: str) -> bool:
    """Check whether the file downloaded has been processed before and insert
    the new file's hash into the database

    Args:
        file_hash (str): An md5 hash of the file from gen_file_hash()

    Returns:
        bool: Whether the file is new or not
    """

    cursor = con.cursor()
    result = cursor.execute(SELECT_FILES_QUERY,
                            (file_hash,)).fetchall()
    print(result)
    print(not result)

    timestamp = int(unix_timestamp())

    cursor.execute(INSERT_FILE_QUERY, (file_hash, timestamp))
    con.commit()

    if len(result) == 0:
        return True
    else:
        return False


def initialise_logs(file_name: str) -> None:
    """Initialise the logs file when program is first run

    Args:
        file_name (str): The name of the file (composed of today's date)
    """
    with open(file_name, "a") as file:
        file.write(constants.LOGGING_SEPARATOR)


def convert_to_xlsx(file_path: str):
    """Convert spreadsheet from xls format to xlsx so that it can be read

    Args:
        file_path (str): the path to the xls file
    """
    import pandas

    df = pandas.read_excel(file_path)

    with pandas.ExcelWriter(
        constants.FILE_DOWNLOAD_LOCATION,
        mode="w",
        engine="openpyxl"
    ) as xlsx:
        df.to_excel(xlsx, index=False)
